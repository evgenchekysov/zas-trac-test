from pathlib import Path
import json
from openpyxl import load_workbook

SRC_PATH = Path("tools/import.xlsx")
OUT_EQUIP = Path("data/equipment.json")
OUT_ROOMS = Path("data/rooms.json")

def norm(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s != "" else None

def main():
    if not SRC_PATH.exists():
        print(f"::warning::No file {SRC_PATH}")
        return

    wb = load_workbook(SRC_PATH, data_only=True)
    ws = wb["Equipment"]

    headers = {cell.value.strip().lower(): idx for idx, cell in enumerate(ws[1], start=1)}

    equipment = []
    rooms = {}

    for row in ws.iter_rows(min_row=2):
        cc = norm(row[headers["cost center"] - 1].value)
        desc = norm(row[headers["asset description"] - 1].value)
        inv = norm(row[headers["inventory number"] - 1].value)
        qr = norm(row[headers["inventory note"] - 1].value)
        room = norm(row[headers["room"] - 1].value)

        if not (desc or inv or qr):
            continue

        equipment.append({
            "cost_center": cc,
            "description": desc,
            "inventory_number": inv,
            "qr_id": qr,
            "room": room
        })

        if room:
            rooms[room] = rooms.get(room, 0) + 1

    OUT_EQUIP.parent.mkdir(parents=True, exist_ok=True)
    OUT_ROOMS.parent.mkdir(parents=True, exist_ok=True)

    with OUT_EQUIP.open("w", encoding="utf-8") as f:
        json.dump(equipment, f, ensure_ascii=False, indent=2)

    rooms_list = [{"room": r, "count": rooms[r]} for r in sorted(rooms)]
    with OUT_ROOMS.open("w", encoding="utf-8") as f:
        json.dump(rooms_list, f, ensure_ascii=False, indent=2)

    print(f"Imported {len(equipment)} items")

if __name__ == "__main__":
    main()
